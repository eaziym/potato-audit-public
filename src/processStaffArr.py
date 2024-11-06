import pandas as pd
from datetime import datetime
from fuzzywuzzy import process, fuzz
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side
import re
from openpyxl.styles import Alignment, Font, NamedStyle
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# def parse_client_names(value):
#     # Split by slashes outside of brackets
#     parts = re.split(r'(?<!\(.*)/(?!\S*\))', value)
#     parsed_names = []
#     for part in parts:
#         # Check if 'Stocktake' is in the part
#         if 'Stocktake' in part:
#             # Extract the name before 'Stocktake' and any dates if present
#             stocktake_info = re.match(r"(.+?) Stocktake(?: - (.+))?", part)
#             if stocktake_info:
#                 client_name = stocktake_info.group(1).strip()
#                 dates = stocktake_info.group(2)
#                 # Parse the dates if they exist
#                 if dates:
#                     date_list = re.findall(r'(\d{1,2}/\d{1,2}/\d{2})', dates)
#                     num_days = len(date_list)  # The number of dates will give us the number of stocktake days
#                 else:
#                     num_days = 5  # Default to a week's duration if no dates are specified
#                 parsed_names.append((client_name, num_days))
#             else:
#                 # If Stocktake is present but the regex didn't match, we fallback to a default
#                 client_name = part.split(' Stocktake')[0].strip()
#                 parsed_names.append((client_name, 5))  # Default to a week's duration
#         else:
#             # Remove content within brackets and trim whitespace
#             client_name = re.sub(r"\s*\([^)]+\)", "", part).strip()
#             parsed_names.append((client_name, None))  # No stocktake, no specific duration
#     return parsed_names


def parse_client_names(value):
    # Define abbreviations to ignore
    ignore_abbreviations = {'AL', 'ML', 'TOIL', 'PHIL', 'Leave', 'Reservist', 'TIL', 'OIL'}

    # Manually split by slashes, taking care not to split within parentheses
    parts = []
    bracket_level = 0
    start = 0
    for i, char in enumerate(value):
        if char == '(':
            bracket_level += 1
        elif char == ')':
            bracket_level -= 1
        elif char == '/' and bracket_level == 0:
            parts.append(value[start:i])
            start = i + 1
    parts.append(value[start:])  # add the last part after the last slash

    # Filter parts: remove content within parentheses, skip stocktakes, dates, and abbreviations
    parsed_names = []
    for part in parts:
        # Remove content within parentheses
        part = re.sub(r"\s*\([^)]+\)", "", part)
        # Skip parts with 'Stocktake', a date, or any listed abbreviation
        if 'Stocktake' in part or 'stocktake' in part or any(abbr in part for abbr in ignore_abbreviations):
            continue
        if re.search(r'\b\d{1,2}/\d{1,2}/\d{2}\b', part):
            continue
        # Trim whitespace and append if not empty
        client_name = part.strip()
        if client_name:
            parsed_names.append(client_name)

    return parsed_names




# Sample data loading - replace this with loading from an actual Excel file
df = pd.read_excel(r'schedules\Schedule 27 May.xlsx', sheet_name='2024')

# Convert date columns assuming they are of type datetime.datetime
date_columns = [col for col in df.columns if isinstance(col, datetime)]

# Level to hourly rate mapping
level_to_rate = {
    'AA': 180,
    'EA': 220,
    'SA1': 250,
    'SA2': 300,
    'AM': 360
}

# Initialize an empty dictionary for client-staff mapping
client_staff_mapping = {}

# Iterate over the DataFrame
for index, row in df.iterrows():
    staff_name = row['Staff Name']
    level = re.match(r"([a-zA-Z]+\d*)", row['Level']).group() if pd.notna(row['Level']) else None

    for date_col in date_columns:
        value = row[date_col]
        if pd.notna(value):
            clients = parse_client_names(str(value))
            date_str = date_col.strftime('%d-%b')
            for client_name in clients:
                if client_name:  # Ensure it's not empty
                    # Check if this client is already recorded for this staff member
                    if client_name not in client_staff_mapping:
                        client_staff_mapping[client_name] = {}
                    if staff_name not in client_staff_mapping[client_name]:
                        client_staff_mapping[client_name][staff_name] = {'weeks': [], 'level': level}
                    # Append the date string to the staff's list for this client
                    client_staff_mapping[client_name][staff_name]['weeks'].append(date_str)

# Convert the nested dictionary to a list of dictionaries, one per client-staff-week mapping
output_data = []
for client, staff_info in client_staff_mapping.items():
    for staff, details in staff_info.items():
        weeks = details['weeks']
        if weeks:  # Make sure there is at least one week present
            week_dates = [datetime.strptime(wk, '%d-%b') for wk in weeks]
            start_date = min(week_dates)
            end_date = max(week_dates)
            date_range = f"{start_date.strftime('%d-%b')} - {end_date.strftime('%d-%b')}"
            level = details['level']
            hourly_rate = level_to_rate.get(level, 0)
            total_hours = 8 * 5 * len(weeks)  # Assuming 40 hours per week
            total_rate = hourly_rate * total_hours

            output_data.append({
                'Client': client, 
                'Staff Assigned': staff, 
                'Weeks': len(weeks), 
                'Level': level,
                'Total Rate': total_rate,
                'Date Range': date_range
            })

# Convert the list to a DataFrame
output_df = pd.DataFrame(output_data)

# Apply fuzzy matching to consolidate similar client names
unique_clients = output_df['Client'].unique()
matches = {client: process.extractOne(client, unique_clients, scorer=fuzz.token_sort_ratio)[0] for client in unique_clients if process.extractOne(client, unique_clients, scorer=fuzz.token_sort_ratio)[1] > 90}
output_df['Consolidated Client'] = output_df['Client'].apply(lambda x: matches.get(x, x))

# Recalculate the job total rates for each consolidated client
consolidated_job_rates = output_df.groupby('Consolidated Client')['Total Rate'].sum().to_dict()

# Map the recalculated job total rates back to the DataFrame
output_df['Job Total Rate'] = output_df['Consolidated Client'].map(consolidated_job_rates)

# def abbreviation_stands_alone(client_name):
#     pattern = r'\b(AL|PHIL|TOIL|Leave|Reservist|ML|TIL|OIL)\b'
#     return bool(re.search(pattern, client_name))

# Replace the 'Client' column with 'Consolidated Client' and drop the latter
output_df['Client'] = output_df['Consolidated Client']
output_df.drop(columns=['Consolidated Client'], inplace=True)

# Ensure the 'Client' column is the first column if not already
output_df = output_df[['Client', 'Staff Assigned', 'Weeks', 'Level', 'Total Rate', 'Date Range', 'Job Total Rate']]

# Filter, sort, and prepare the DataFrame as before
output_df = output_df[output_df['Client'].str.strip().astype(bool)]
# output_df = output_df[~output_df['Client'].apply(abbreviation_stands_alone)]
output_df.sort_values(by=['Client', 'Staff Assigned'], inplace=True)

# Sort the DataFrame by 'Job Total Rate' in descending order
sorted_by_rate_df = output_df.sort_values(by=['Job Total Rate', 'Client'], ascending=False)



def apply_styles_and_borders(sheet, df):
    header_font = Font(bold=True)
    client_font = Font(bold=True)
    alignment_center = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = header_font
                cell.alignment = alignment_center
                cell.fill = header_fill
            else:
                cell.border = thin_border
                cell.alignment = alignment_center
                if c_idx == 1:  # Client names are in the first column
                    cell.font = client_font
                    cell.alignment = alignment_center

def set_column_widths(sheet):
    for column in sheet.columns:
        max_length = 0
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width




def apply_thick_border_to_group(sheet, start_row, end_row, start_col, end_col):
    thick_border_side = Side(style='thick')
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            cell.border = Border(
                top=thick_border_side if row == start_row else cell.border.top,
                bottom=thick_border_side if row == end_row else cell.border.bottom,
                left=thick_border_side if col == start_col else cell.border.left,
                right=thick_border_side if col == end_col else cell.border.right
            )

def merge_similar_cells(sheet, column_index):
    previous_cell = None
    start_row = None
    for row in range(1, sheet.max_row + 1):
        current_cell = sheet.cell(row=row, column=column_index)
        if previous_cell and current_cell.value == previous_cell.value:
            if start_row is None:
                start_row = previous_cell.row  # Start a new merge range
        else:
            if start_row is not None:
                sheet.merge_cells(start_row=start_row, start_column=column_index, end_row=previous_cell.row, end_column=column_index)
                apply_thick_border_to_group(sheet, start_row, previous_cell.row, column_index, column_index)
                start_row = None  # Reset start_row for the next merge range
        previous_cell = current_cell
    # Check if the last group needs merging
    if start_row is not None:
        sheet.merge_cells(start_row=start_row, start_column=column_index, end_row=sheet.max_row, end_column=column_index)
        apply_thick_border_to_group(sheet, start_row, sheet.max_row, column_index, column_index)

# Function to check if the name is just numbers and special characters
def is_not_valid_client(name):
    # This regex matches strings that contain only numbers and/or special characters
    return re.match(r'^[\d\W]+$', name) is not None

# Filter the DataFrame
output_df = output_df[~output_df['Client'].apply(is_not_valid_client)]
sorted_by_rate_df = sorted_by_rate_df[~sorted_by_rate_df['Client'].apply(is_not_valid_client)]

# Create a new workbook and select the active worksheet
workbook = Workbook()
main_sheet = workbook.active
main_sheet.title = "Staff Assignments"

# Add another sheet for the jobs sorted by rates
rates_sheet = workbook.create_sheet(title="Jobs by Rates")

# Assuming output_df is already filled with data and sorted_by_rate_df is sorted
apply_styles_and_borders(main_sheet, output_df)
apply_styles_and_borders(rates_sheet, sorted_by_rate_df)

# Set column widths after populating data
set_column_widths(main_sheet)
set_column_widths(rates_sheet)


merge_similar_cells(main_sheet, 1)  # Client name in the first column
merge_similar_cells(main_sheet, 7)  # Job Total Rate in the seventh column

merge_similar_cells(rates_sheet, 1)  # Client name in the first column
merge_similar_cells(rates_sheet, 7)  # Job Total Rate in the seventh column

# Save the workbook
filename = f"Client_Staff_Mapping_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
workbook.save(f"staff_mapping/{filename}")
