import pandas as pd
from openpyxl import load_workbook, Workbook
from datetime import datetime
import os
import numpy as np

EXCEL_FILE_PATH = 'bookings.xlsx'

# # Level to hourly rate mapping
# level_to_rate = {
#     'AA': 180,
#     'EA': 220,
#     'SA1': 250,
#     'SA2': 300,
#     'AM': 360
# }

def workdays(start_date, end_date, holidays=[]):
    # Create a range of dates between start_date and end_date
    date_range = pd.date_range(start=start_date, end=end_date)
    # Filter out the weekends
    weekdays = date_range[date_range.dayofweek < 5]
    # Further filter out any holidays
    working_days = weekdays[~weekdays.isin(holidays)]
    return len(working_days)

def ensure_workbook():
    if os.path.exists(EXCEL_FILE_PATH):
        wb = load_workbook(EXCEL_FILE_PATH)
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])  # Remove the default sheet if it's empty
        return wb
    else:
        wb = Workbook()
        wb.create_sheet("Staff to Client Mapping")
        wb.create_sheet("Client to Staff Mapping")
        wb.remove(wb['Sheet'])
    return wb

def save_to_excel(book):
    book.save(EXCEL_FILE_PATH)
    book.close()

def find_or_create_row(sheet, search_value, headers):
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=2).value == search_value:
            return row
    # Append new row if not found
    new_row = sheet.max_row + 1
    sheet.append([new_row-1] + [''] * (len(headers) - 1))
    return new_row

def update_mappings(book, booking_data, level_to_rate):
    headers_staff = ["Index", "Staff Name", "Level", "Time Cost", "Booking Details"]
    headers_client = ["Index", "Client Info", "Audit Fee", "Total Time Cost", "Recovery Rate", "Staff Details"]

    staff_client_sheet = book.create_sheet("Staff to Client Mapping") if "Staff to Client Mapping" not in book.sheetnames else book["Staff to Client Mapping"]
    client_staff_sheet = book.create_sheet("Client to Staff Mapping") if "Client to Staff Mapping" not in book.sheetnames else book["Client to Staff Mapping"]
    
    # Ensure headers are set
    if staff_client_sheet.max_row == 1:
        staff_client_sheet.append(headers_staff)
    if client_staff_sheet.max_row == 1:
        client_staff_sheet.append(headers_client)

    staff_name = booking_data['staff']['name']
    client_info = f"{booking_data['client']} - {booking_data['financialYear']}"

    row_number_staff = find_or_create_row(staff_client_sheet, staff_name, headers_staff)
    row_number_client = find_or_create_row(client_staff_sheet, client_info, headers_client)

    start_date = datetime.strptime(booking_data['startDate'], '%Y-%m-%d')
    end_date = datetime.strptime(booking_data['endDate'], '%Y-%m-%d')
    # Call the workdays function to get the number of working days
    days_booked = workdays(booking_data['startDate'], booking_data['endDate'], holidays=[])
    hourly_rate = float(level_to_rate.get(booking_data['staff']['level'], 0))
    time_cost = hourly_rate * 8 * days_booked

    booking_details = f"{booking_data['client']} - {booking_data['financialYear']} - {booking_data['startDate']} to {booking_data['endDate']} - ${booking_data['auditFee']}"

    total_time_cost_cell = staff_client_sheet.cell(row=row_number_staff, column=4)
    existing_time_cost = total_time_cost_cell.value or 0
    total_time_cost = float(existing_time_cost) + float(time_cost)

    staff_client_sheet.cell(row=row_number_staff, column=2, value=staff_name)
    staff_client_sheet.cell(row=row_number_staff, column=3, value=booking_data['staff']['level'])
    staff_client_sheet.cell(row=row_number_staff, column=4, value=total_time_cost)
    col_index = 5
    while staff_client_sheet.cell(row=row_number_staff, column=col_index).value:
        col_index += 1
    staff_client_sheet.cell(row=row_number_staff, column=col_index, value=booking_details)


    audit_fee = float(booking_data['auditFee']) if booking_data['auditFee'] else 0
    total_time_cost_cell = client_staff_sheet.cell(row=row_number_client, column=4)
    existing_time_cost = total_time_cost_cell.value or 0
    total_time_cost = float(existing_time_cost) + float(time_cost)
    recovery_rate = (audit_fee / total_time_cost) * 100 if total_time_cost > 0 else 0

    client_staff_sheet.cell(row=row_number_client, column=2, value=client_info)
    client_staff_sheet.cell(row=row_number_client, column=3, value="{:.2f}".format(audit_fee))
    client_staff_sheet.cell(row=row_number_client, column=4, value="{:.2f}".format(total_time_cost))
    client_staff_sheet.cell(row=row_number_client, column=5, value="{:.2f}%".format(recovery_rate))

    col_index = 6
    while client_staff_sheet.cell(row=row_number_client, column=col_index).value:
        col_index += 1
    client_staff_sheet.cell(row=row_number_client, column=col_index, value=f"{staff_name} - {booking_data['staff']['level']} - {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')} - ${time_cost:.2f}")


def process_booking(booking_data, rates):
    book = ensure_workbook()
    update_mappings(book, booking_data, rates)
    save_to_excel(book)
    # Close the workbook
    book.close()
