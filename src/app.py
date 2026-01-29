from flask import Flask, jsonify, request, send_file, send_from_directory
from urllib.parse import unquote
import os
from GetSGX_Announcement import fetch_SGX_announcements  
from GetASX_Announcement import fetch_ASX_announcements  
from GetSGXAnnouncement_Today import fetch_announcements_for_companies
from emailANN import emailANN
from GetSGX_KAM import fetch_and_process_annual_reports  # Updated import
from ProcessPDF import process_pdfs_and_generate_excel
from ProcessDocs import process_docs_and_generate_excel
import cv2
from werkzeug.utils import secure_filename
from process_invoice_file import process_invoice_image
from flask_cors import CORS
from process_pairs import get_spatial_relationship
from process_pairs import apply_spatial_relationship
from process_pairs import find_coordinates_by_text
from process_pairs import find_closest_box
from process_pairs import sanitize_text
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import json
import time
import datetime
import pandas as pd
import re
from excel_utils import process_booking



app = Flask(__name__)
CORS(app)

@app.route('/fetch-process-reports', methods=['POST'])
def fetch_process_reports():
    data = request.get_json()
    download_dir = data.get('downloadDir')

    # Validate downloadDir
    if not download_dir or not os.path.isdir(download_dir):
        return jsonify({"error": "Invalid or missing download directory."}), 400

    try:
        # Call the function to fetch and process the reports
        fetch_and_process_annual_reports(download_dir)
        return jsonify({"success": True, "message": "Reports fetched and processed successfully."})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/fetch-announcements', methods=['POST'])
def handle_fetch_announcements():
    data = request.get_json()
    company = data.get('company')
    year = data.get('year')
    exchange = data.get('exchange')
    downloadDir = data.get('downloadDir')

    # Ensure the download directory exists and is accessible
    target_dir = os.path.join(downloadDir, f"{company.replace(' ', '_').upper()}_FY{year}")
    os.makedirs(target_dir, exist_ok=True)

    try:
        # Step 1: Fetch and Download Announcements based on the exchange
        if exchange == 'SGX':
            fetch_SGX_announcements(company, year, target_dir)
        elif exchange == 'ASX':
            fetch_ASX_announcements(company, year, target_dir)
        else:
            return jsonify({"error": f"Exchange {exchange} not supported"}), 400


        print("Starting PDF processing...")
        try:
            # Step 2: Process PDFs and Generate Excel
            excel_path = process_pdfs_and_generate_excel(target_dir)
            
            # Serve the Excel file directly to the user
            return send_file(excel_path, as_attachment=True, download_name=os.path.basename(excel_path))
        except Exception as e:
            print(f"Error during PDF processing: {e}")
            raise  # Re-raise the exception to catch it in the Flask error handler

    except Exception as e:
        return jsonify({"error": str(e)[:20]}), 500
    

@app.route('/fetch-todays-announcements', methods=['POST'])
def fetch_todays_announcements():
    data = request.get_json()
    companies = data.get('companies', '').split(',')
    downloadDir = data.get('downloadDir')

    
    results = []
    
    # Fetch announcements for all companies in a single call
    results = fetch_announcements_for_companies(companies)

    # Save results to CSV
    yesterday = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime("%Y-%m-%d")
    filename = f"{yesterday}_announcements.csv"
    filepath = os.path.join(downloadDir, filename)  # Save the file in the provided download directory
    with open(filepath, 'w', newline='') as csvfile:
        fieldnames = ['date', 'company', 'title', 'link']
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        
        writer.writeheader()
        for row in results:
            writer.writerow(row)

    # Check if there are any announcements to email
    # if the column 'title' contains all 'No announcements today.', then no announcements to email
    check = all(row['title'] == 'No announcements today.' for row in results)
    if check==False:
        emailANN(filepath)
    
    
    return jsonify({'status': 'success', 'message': f'Announcements saved to {filename}'})


UPLOAD_FOLDER = 'uploads'
PROCESSED_FOLDER = 'processed_invoices'

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['PROCESSED_FOLDER'] = PROCESSED_FOLDER

progress = {'current': 0, 'total': 1}  # Initialize progress


@app.route('/process-directory', methods=['POST'])
def process_directory():
    data = request.json
    directory_path = data.get('directory')
    
    # Validate the directory path here (ensure it's a safe path and accessible)
    
    # Find all invoice files in the directory (filter by .pdf and .png)
    invoice_files = [f for f in os.listdir(directory_path) if f.lower().endswith(('.pdf', '.png'))]
    
    if not invoice_files:
        return jsonify({"error": "No invoice files found in the specified directory."}), 404

    # Process the first invoice file as a sample
    sample_invoice_path = os.path.join(directory_path, invoice_files[0])
    # Assume process_invoice returns data necessary for creating a template
    processed_path, blobs, mergedTexts = process_invoice_image(sample_invoice_path, PROCESSED_FOLDER)
    filename = os.path.basename(processed_path)    
    # Send back information to allow the user to create a template
    # For simplicity, this might just acknowledge the process has started
    return jsonify({"status": "success", "message": "Sample invoice processed. Ready for template creation.", "filename": filename, "blobs": blobs, "mergedTexts": mergedTexts} )

# Static files serving route for processed images
@app.route('/processed-images/<filename>')
def serve_processed_image(filename):
    return send_from_directory(app.config['PROCESSED_FOLDER'], filename)


@app.route('/upload-docs', methods=['POST'])
def upload_docs():
    uploaded_files = request.files.getlist('files')
    download_directory = request.form.get('downloadDirectory')


    # Ensure the user-specified directory exists
    if not os.path.exists(download_directory):
        os.makedirs(download_directory)

    # Save uploaded files to the general uploads directory
    for file in uploaded_files:
        if file:
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)

    # Process the files from the upload directory and save the output in the specified download directory
    try:
        excel_path = process_docs_and_generate_excel(UPLOAD_FOLDER, download_directory)
        return jsonify({"success": True, "message": f"Great! File processed successfully and saved to {excel_path}"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500



@app.route('/process-selected-pairs', methods=['POST'])
def process_selected_pairs():
    data = request.json
    selected_pairs = data.get('selectedPairs')
    blobs = data.get('blobs')
    merged_texts = data.get('mergedTexts')
    directoryPath = data.get('directoryPath')
    labelVariations = data.get('labelVariations', {})
    print(labelVariations)


    # Validate the essential data is provided
    if not all([selected_pairs, blobs, merged_texts]):
        return jsonify({"error": "Missing data in the request"}), 400

    def find_text_by_coordinates(target_coords, merged_texts):
        for item in merged_texts:
            if item['coordinates'] == target_coords:
                return item['text']
        return "Text not found"

    final_output = {}
    spatial_relationships = {}

    for index, (label_index, value_index) in enumerate(selected_pairs):
        label_coords = blobs[label_index]['coordinates']
        value_coords = blobs[value_index]['coordinates']

        # Find corresponding texts for the given coordinates
        label_text = find_text_by_coordinates(label_coords, merged_texts)
        value_text = find_text_by_coordinates(value_coords, merged_texts)
        dx, dy = get_spatial_relationship(label_coords, value_coords)
        spatial_relationships[label_text] = (dx, dy)

        final_output[index] = {
            'label_coords': label_coords,
            'value_coords': value_coords,
            'label_text': label_text,
            'value_text': value_text,
            'spatial_relationship': (dx, dy)
        }

    # Track progress
    invoice_files = [f for f in os.listdir(directoryPath) if f.lower().endswith(('.pdf', '.png'))]
    progress['total'] = len(invoice_files)
    progress['current'] = 0

    processed_info = []  # Store processed invoice info here
    for index, invoice_file in enumerate(invoice_files, start=1):
        print(f"Processing {invoice_file}...")
        progress['current'] = index
        invoice_path = os.path.join(directoryPath, invoice_file)
        processed_path, invoice_blobs, invoice_merged_texts = process_invoice_image(invoice_path, PROCESSED_FOLDER)
        print(invoice_merged_texts)
        invoice_data = {}  # Dictionary to store extracted data from this invoice
    
        # Iterate through selected pairs and extract relevant data
        for label_text, (dx, dy) in spatial_relationships.items():  # Assuming spatial_relationships is available
            # Find label box in the current invoice using label_text
            print('Current label: ' + label_text)
            label_box, found_label_text = find_coordinates_by_text(label_text, invoice_merged_texts, dx, dy, labelVariations, threshold=90)

            if label_box is None:
                print('Label box not found for: ' + label_text)
                invoice_data[label_text] = None
                continue
            print('Label box found: ' + str(label_box))
            print('Found label box with text: ' + found_label_text)

            # Apply spatial relationship to estimate value coordinates
            estimated_value_coords = apply_spatial_relationship(label_box, dx, dy)
            

            # Find the closest box to the estimated value coordinates
            closest_box = find_closest_box(dx, dy, estimated_value_coords, invoice_merged_texts, omit_box=label_box)
            if closest_box:
                value_text = closest_box['text']
                if dx == 0 and dy == 0:
                    label_text = sanitize_text(label_text)
                invoice_data[label_text] = value_text
                print(label_text, value_text)
            elif closest_box is None:
                print('Closest box not found for: ' + found_label_text)
                print('Try with second best match...')
                label_box, found_label_text = find_coordinates_by_text(label_text, invoice_merged_texts, dx, dy, labelVariations, threshold=90, omit_text=found_label_text)
                if label_box is None:
                    print('Second best match also not found for: ' + label_text)
                    invoice_data[label_text] = None
                    continue
                print('Second best match found: ' + str(label_box))
                print('Found second best match with text: ' + found_label_text)
                estimated_value_coords = apply_spatial_relationship(label_box, dx, dy)
                closest_box = find_closest_box(dx, dy, estimated_value_coords, invoice_merged_texts)
                if closest_box:
                    value_text = closest_box['text']
                    if dx == 0 and dy == 0:
                        label_text = sanitize_text(label_text)
                    invoice_data[label_text] = value_text
                    print(label_text, value_text)
                elif closest_box is None:
                    print('Closest box not found for second best match: ' + label_text)
                    invoice_data[label_text] = None

        # Add the file URL to the processed data
        invoice_data['File URL'] = invoice_path
        processed_info.append(invoice_data)
        print(processed_info[0].keys())

    wb = Workbook()
    ws = wb.active

    # Adding headers
    headers = processed_info[0].keys()
    for col, header in enumerate(headers, start=1):
        ws[f"{get_column_letter(col)}1"] = header

    # Adding data rows
    for row, info in enumerate(processed_info, start=2):
        for col, (key, value) in enumerate(info.items(), start=1):
            if key == 'File URL':
                # Insert a HYPERLINK formula for the file path
                ws[f"{get_column_letter(col)}{row}"] = f'=HYPERLINK("{value}", "Click Here to Open File")'
            else:
                if value:
                    ws[f"{get_column_letter(col)}{row}"] = value

    output_filename = os.path.join(directoryPath, "processed_invoices.xlsx")  # Save in the user's directory
    wb.save(output_filename)


    output_json_filename = os.path.join(directoryPath, "processed_info.json")
    with open(output_json_filename, 'w') as jsonfile:
        json.dump(processed_info, jsonfile, indent=4)

    return jsonify({"status": "success", "output": final_output, "spatial_relationships": spatial_relationships})


@app.route('/get-processed-info', methods=['GET'])
def get_processed_info():
    directory_path = request.args.get('directory')  # Assume directory path is provided as a query parameter
    directory_path = unquote(directory_path)
    if not directory_path:
        return jsonify({"error": "No directory path provided."}), 400

    # Path to the file where processed invoice data is stored
    data_file_path = os.path.join(directory_path, 'processed_info.json')
   
    attempts = 0
    max_attempts = 10
    while attempts < max_attempts:
        try:
            with open(data_file_path, 'r') as file:
                data = json.load(file)
                return jsonify(data)
        except IOError as e:
            app.logger.error(f"An error occurred: {str(e)}")
            time.sleep(2)  # Wait for 1 second before retrying
            attempts += 1


@app.route('/progress')
def get_progress():
    return jsonify(progress)


@app.route('/api/staff', methods=['GET'])
def get_staff():
    # Make sure to provide the correct path to your Excel file
    try:
        # Assuming the staff names are in column 'B' and levels in column 'C'
        df = pd.read_excel('staff_names.xlsx', sheet_name='Sheet1', usecols="B:C", skiprows=0, nrows=620)

        # Clean and process the 'Level' data
        df['Level'] = df['Level'].apply(lambda x: re.sub(r"\s*\([^)]+\)", "", str(x)).strip())

        # Create a list of dictionaries with staff names and their levels
        staff_data = [{
            'name': row["Staff's Name"],
            'level': row['Level']
        } for index, row in df.dropna().iterrows()]  # Drop any rows where data may be missing

        return jsonify({'staff': staff_data})
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/clients', methods=['GET'])
def get_clients():
    # Make sure to provide the correct path to your Excel file
    try:
        df = pd.read_excel('client_names.xlsx', sheet_name='Sheet1', usecols='B', skiprows=0, nrows=620)

        client_names = df["Client's Name"].dropna().tolist()        # Convert the data to a list, removing any NaN values
        unique_client_names = list(set(client_names))
        # Sort the list if needed, as sets do not retain order
        unique_client_names.sort()
        return jsonify({'clients': unique_client_names})
    except Exception as e:
        return jsonify({'error': str(e)})


level_to_rate = {
    'AA': 180,
    'EA': 220,
    'SA1': 250,
    'SA2': 300,
    'AM': 360
}

@app.route('/api/bookings', methods=['POST'])
def create_booking():
    booking_data = request.json
    # Print the booking data to the console
    rates = booking_data.get('rates', level_to_rate)
    print(booking_data)
    process_booking(booking_data, rates)  # Call the function to process and update Excel
    return jsonify({"status": "success", "message": "Booking processed and saved to Excel."})

@app.route('/')
def home():
    return 'Hello, World!'

if __name__ == '__main__':
    # Create necessary folders if they don't exist
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    os.makedirs(app.config['PROCESSED_FOLDER'], exist_ok=True)
    
    app.run(debug=True, port=5002)



